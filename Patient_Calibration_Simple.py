"""
SIMPLIFIED ROM CALIBRATION
==========================
Calibrates patient by doing ONE repetition of each exercise in their training session.
Reuses existing exercise infrastructure (audio, robot, angle tracking).

Author: Simplified version - November 2025
"""

import time
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import Settings as s


class Simple_Calibration:
    """
    Simplified calibration: Patient does 1 rep of each exercise as warm-up
    Records max/min ranges during those movements
    """
    
    EXCEL_FILE = "PatientROM_Simple.xlsx"
    
    def __init__(self):
        """Initialize simplified calibration"""
        self.ensure_excel_exists()
    
    def ensure_excel_exists(self):
        """Create Excel file if it doesn't exist"""
        if not os.path.exists(self.EXCEL_FILE):
            print(f"📂 Creating calibration database: {self.EXCEL_FILE}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Calibration_Data"
            
            # Simple headers: PatientID, Date, Exercise, Right_Max, Right_Min, Left_Max, Left_Min
            headers = ['PatientID', 'Date', 'Exercise', 
                      'Right_Max', 'Right_Min', 'Left_Max', 'Left_Min']
            ws.append(headers)
            wb.save(self.EXCEL_FILE)
            print(f"✅ Created {self.EXCEL_FILE}")
    
    def run_calibration_for_training(self):
        """
        Run calibration for current training session
        Patient does 1 rep of each exercise in s.ex_in_training
        """
        if not hasattr(s, 'ex_in_training') or not s.ex_in_training:
            print("❌ No exercises in training session!")
            return False
        
        print("\n" + "="*70)
        print(f"🎯 WARM-UP CALIBRATION")
        print(f"   Patient: {s.chosen_patient_ID}")
        print(f"   Exercises in session: {len(s.ex_in_training)}")
        print("="*70)
        print("\n📋 You will do ONE repetition of each exercise.")
        print("   This helps us personalize your workout!\n")
        
        calibration_data = {}
        
        # For each exercise in training session
        for idx, exercise in enumerate(s.ex_in_training, 1):
            if s.stop_requested:
                print("\n⚠️ Calibration cancelled")
                return False
            
            print(f"\n[{idx}/{len(s.ex_in_training)}] Calibrating: {exercise}")
            print(f"   🤖 Watch the robot, then do ONE repetition...")
            
            # Update GUI
            s.current_calibration_movement = f"{exercise} (1 repetition)"
            s.current_calibration_progress = f"{idx}/{len(s.ex_in_training)}"
            
            # Do one calibration repetition of this exercise
            ranges = self.calibrate_one_exercise(exercise)
            
            if ranges:
                calibration_data[exercise] = ranges
                print(f"   ✅ Recorded ranges for {exercise}")
            else:
                print(f"   ⚠️ Skipped {exercise}")
        
        # Save to Excel
        self.save_to_excel(s.chosen_patient_ID, calibration_data)
        
        # Store in Settings for this session
        s.patient_rom = calibration_data
        s.patient_calibrated = True
        
        print("\n" + "="*70)
        print("✅ WARM-UP CALIBRATION COMPLETE!")
        print("   Let's start your training!")
        print("="*70)
        
        return True
    
    def calibrate_one_exercise(self, exercise_name):
        """
        Do ONE repetition of an exercise to calibrate ranges
        
        Returns:
            dict: {'right_max': angle, 'right_min': angle, 
                   'left_max': angle, 'left_min': angle}
        """
        # Set up for calibration mode
        s.req_exercise = exercise_name
        s.patient_repetitions_counting_in_exercise = 0
        s.calibration_mode = True  # New flag to tell Camera to track ranges
        s.calibration_ranges = {'right_max': 0, 'right_min': 180,
                                'left_max': 0, 'left_min': 180}
        
        # Wait for robot to demonstrate (existing system handles this)
        time.sleep(2)  # Brief pause for robot demo
        
        # Wait for patient to complete ONE repetition
        # Camera will update s.calibration_ranges during the rep
        timeout = 30  # 30 seconds to do one rep
        start_time = time.time()
        
        while s.patient_repetitions_counting_in_exercise < 1:
            if time.time() - start_time > timeout:
                print(f"   ⏱️ Timeout for {exercise_name}")
                break
            
            if s.stop_requested:
                break
            
            time.sleep(0.1)
        
        # Stop this exercise
        s.req_exercise = ""
        s.calibration_mode = False
        
        # Return the ranges
        if s.patient_repetitions_counting_in_exercise >= 1:
            return {
                'right_max': s.calibration_ranges['right_max'],
                'right_min': s.calibration_ranges['right_min'],
                'left_max': s.calibration_ranges['left_max'],
                'left_min': s.calibration_ranges['left_min']
            }
        else:
            return None
    
    def save_to_excel(self, patient_id, calibration_data):
        """Save calibration data to Excel"""
        try:
            wb = load_workbook(self.EXCEL_FILE)
            ws = wb.active
            
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            # Save one row per exercise
            for exercise, ranges in calibration_data.items():
                row = [
                    patient_id,
                    date_str,
                    exercise,
                    ranges.get('right_max', 0),
                    ranges.get('right_min', 0),
                    ranges.get('left_max', 0),
                    ranges.get('left_min', 0)
                ]
                ws.append(row)
            
            wb.save(self.EXCEL_FILE)
            print(f"\n💾 Calibration saved to: {self.EXCEL_FILE}")
            return True
            
        except Exception as e:
            print(f"❌ Error saving: {e}")
            return False
    
    def load_from_excel(self, patient_id):
        """Load most recent calibration for patient"""
        try:
            if not os.path.exists(self.EXCEL_FILE):
                return None
            
            df = pd.read_excel(self.EXCEL_FILE)
            patient_data = df[df['PatientID'] == patient_id]
            
            if patient_data.empty:
                return None
            
            # Get most recent session (by date)
            patient_data = patient_data.sort_values('Date', ascending=False)
            
            # Convert to dict format
            calibration = {}
            for _, row in patient_data.iterrows():
                exercise = row['Exercise']
                calibration[exercise] = {
                    'right_max': row['Right_Max'],
                    'right_min': row['Right_Min'],
                    'left_max': row['Left_Max'],
                    'left_min': row['Left_Min']
                }
            
            return calibration
            
        except Exception as e:
            print(f"⚠️ Error loading calibration: {e}")
            return None


# Helper function for easy import
def run_simple_calibration():
    """Simple function to run calibration"""
    calibration = Simple_Calibration()
    return calibration.run_calibration_for_training()

