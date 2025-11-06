"""
Patient_Calibration_Complete.py - Comprehensive ROM Calibration System
- Measures ALL joint angles needed for exercises
- Adjusted for SEATED position (realistic ROM values)
- Saves to Excel (PatientROM_Calibration.xlsx)
- Loads from Excel when patient starts training
- Provides adaptive ranges for ALL exercises
"""

import time
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import Settings as s
from Audio import say, get_wav_duration

# Audio file mapping for calibration instructions
# TODO: Record these audio files and place in audio_files/Hebrew/Male(Female)/
CALIBRATION_AUDIO = {
    'start': 'calibration_start',  # "Let's measure your range of motion"
    'shoulder_forward': 'calibration_shoulder_forward',  # "Raise your arm forward and up"
    'shoulder_side': 'calibration_shoulder_side',  # "Raise your arm to the side"
    'elbow_bend': 'calibration_elbow_bend',  # "Bend your elbow"
    'arm_across': 'calibration_arm_across',  # "Move your arm across your body"
    'hold_position': 'calibration_hold_position',  # "Hold this position"
    'return_rest': 'calibration_return_rest',  # "Return to rest position"
    'good_job': 'calibration_good_job',  # "Good job"
    'next_movement': 'calibration_next_movement',  # "Next movement"
    'complete': 'calibration_complete',  # "Calibration complete!"
}


class Patient_Calibration:
    """
    Comprehensive ROM Calibration System for SEATED Exercises
    Measures all angles needed for your 24 exercises
    Adjusted ROM values for seated position
    """
    
    EXCEL_FILE = "PatientROM_Calibration.xlsx"
    
    def __init__(self):
        """Initialize calibration system"""
        
        # Comprehensive calibration measurements
        # Based on analysis of ALL 24 exercises
        # MUST be defined BEFORE ensure_excel_exists() is called
        self.calibration_measurements = [
            # ==== SHOULDER FLEXION (arm forward/up) - SEATED ====
            {
                'name': 'R_Shoulder_Hip_Elbow',
                'display': 'Right Shoulder - Raise Arm Forward',
                'joints': ('R_hip', 'R_shoulder', 'R_elbow'),
                'instruction': 'Raise your RIGHT arm FORWARD and UP as high as you can',
                'rest_instruction': 'Now lower your RIGHT arm down to your side',
                'normal_max': 180,
                'normal_min': 10,
                'used_in': ['ball_raise_arms_above_head', 'stick_raise_arms_above_head', 'band_open_arms_and_up']
            },
            {
                'name': 'L_Shoulder_Hip_Elbow',
                'display': 'Left Shoulder - Raise Arm Forward',
                'joints': ('L_hip', 'L_shoulder', 'L_elbow'),
                'instruction': 'Raise your LEFT arm FORWARD and UP as high as you can',
                'rest_instruction': 'Now lower your LEFT arm down to your side',
                'normal_max': 180,
                'normal_min': 10,
                'used_in': ['ball_raise_arms_above_head', 'stick_raise_arms_above_head']
            },
            
            # ==== SHOULDER ABDUCTION (arm to side) ====
            {
                'name': 'R_Shoulder_Hip_Wrist',
                'display': 'Right Shoulder - Raise Arm to Side',
                'joints': ('R_hip', 'R_shoulder', 'R_wrist'),
                'instruction': 'Raise your RIGHT arm OUT TO THE SIDE as high as you can',
                'rest_instruction': 'Now lower your RIGHT arm to your side',
                'normal_max': 150,
                'normal_min': 20,
                'used_in': ['band_open_arms', 'notool_right_hand_up_and_bend', 'weights_abduction']
            },
            {
                'name': 'L_Shoulder_Hip_Wrist',
                'display': 'Left Shoulder - Raise Arm to Side',
                'joints': ('L_hip', 'L_shoulder', 'L_wrist'),
                'instruction': 'Raise your LEFT arm OUT TO THE SIDE as high as you can',
                'rest_instruction': 'Now lower your LEFT arm to your side',
                'normal_max': 150,
                'normal_min': 20,
                'used_in': ['band_open_arms', 'notool_left_hand_up_and_bend']
            },
            
            # ==== ELBOW FLEXION (bending elbow) ====
            {
                'name': 'R_Elbow',
                'display': 'Right Elbow - Bend',
                'joints': ('R_shoulder', 'R_elbow', 'R_wrist'),
                'instruction': 'BEND your RIGHT elbow, bring hand to shoulder',
                'rest_instruction': 'Now STRAIGHTEN your RIGHT elbow completely',
                'normal_max': 150,
                'normal_min': 5,
                'used_in': ['ball_bend_elbows', 'stick_bend_elbows', 'all elbow exercises']
            },
            {
                'name': 'L_Elbow',
                'display': 'Left Elbow - Bend',
                'joints': ('L_shoulder', 'L_elbow', 'L_wrist'),
                'instruction': 'BEND your LEFT elbow, bring hand to shoulder',
                'rest_instruction': 'Now STRAIGHTEN your LEFT elbow completely',
                'normal_max': 150,
                'normal_min': 5,
                'used_in': ['ball_bend_elbows', 'stick_bend_elbows', 'all elbow exercises']
            },
            
            # ==== SHOULDER ROTATION (elbow-shoulder-hip angle) ====
            {
                'name': 'R_Elbow_Shoulder_Hip',
                'display': 'Right Shoulder - Arm Away from Body',
                'joints': ('R_elbow', 'R_shoulder', 'R_hip'),
                'instruction': 'Raise your RIGHT elbow OUT TO THE SIDE',
                'rest_instruction': 'Now bring your RIGHT elbow back to your side',
                'normal_max': 180,
                'normal_min': 0,
                'used_in': ['ball_open_arms_above_head', 'band_up_and_lean']
            },
            {
                'name': 'L_Elbow_Shoulder_Hip',
                'display': 'Left Shoulder - Arm Away from Body',
                'joints': ('L_elbow', 'L_shoulder', 'L_hip'),
                'instruction': 'Raise your LEFT elbow OUT TO THE SIDE',
                'rest_instruction': 'Now bring your LEFT elbow back to your side',
                'normal_max': 180,
                'normal_min': 0,
                'used_in': ['ball_open_arms_above_head', 'band_up_and_lean']
            },
            
            # ==== SHOULDER HORIZONTAL ADDUCTION (wrist-shoulder-shoulder) ====
            {
                'name': 'R_Wrist_Shoulder_Shoulder',
                'display': 'Right Arm - Across Body',
                'joints': ('R_wrist', 'R_shoulder', 'L_shoulder'),
                'instruction': 'Bring your RIGHT hand across your body to the LEFT',
                'rest_instruction': 'Now bring your RIGHT hand back out to the side',
                'normal_max': 180,
                'normal_min': 60,
                'used_in': ['ball_open_arms_and_forward', 'band_open_arms', 'weights_open_arms_and_forward']
            },
            {
                'name': 'L_Wrist_Shoulder_Shoulder',
                'display': 'Left Arm - Across Body',
                'joints': ('L_wrist', 'L_shoulder', 'R_shoulder'),
                'instruction': 'Bring your LEFT hand across your body to the RIGHT',
                'rest_instruction': 'Now bring your LEFT hand back out to the side',
                'normal_max': 180,
                'normal_min': 60,
                'used_in': ['ball_open_arms_and_forward', 'band_open_arms']
            },
            
            # ==== SHOULDER EXTENSION (wrist-hip-hip angle for switch exercises) ====
            {
                'name': 'R_Wrist_Hip_Hip',
                'display': 'Right Arm - Behind Body',
                'joints': ('R_wrist', 'R_hip', 'L_hip'),
                'instruction': 'Move your RIGHT hand behind your back',
                'rest_instruction': 'Bring your RIGHT hand back to front',
                'normal_max': 160,
                'normal_min': 35,
                'used_in': ['ball_switch', 'stick_switch', 'band_up_and_lean']
            },
            {
                'name': 'L_Wrist_Hip_Hip',
                'display': 'Left Arm - Behind Body',
                'joints': ('L_wrist', 'L_hip', 'R_hip'),
                'instruction': 'Move your LEFT hand behind your back',
                'rest_instruction': 'Bring your LEFT hand back to front',
                'normal_max': 160,
                'normal_min': 35,
                'used_in': ['ball_switch', 'stick_switch']
            },
            
            # ==== REVERSE ANGLES (wrist-elbow-shoulder for certain exercises) ====
            {
                'name': 'R_Wrist_Elbow_Shoulder',
                'display': 'Right Arm - Straightness',
                'joints': ('R_wrist', 'R_elbow', 'R_shoulder'),
                'instruction': 'STRAIGHTEN your RIGHT arm completely',
                'rest_instruction': 'Relax your RIGHT arm',
                'normal_max': 180,
                'normal_min': 120,
                'used_in': ['stick_raise_arms_above_head', 'weights_abduction', 'notool_right_bend_left_up_from_side']
            },
            {
                'name': 'L_Wrist_Elbow_Shoulder',
                'display': 'Left Arm - Straightness',
                'joints': ('L_wrist', 'L_elbow', 'L_shoulder'),
                'instruction': 'STRAIGHTEN your LEFT arm completely',
                'rest_instruction': 'Relax your LEFT arm',
                'normal_max': 180,
                'normal_min': 120,
                'used_in': ['stick_raise_arms_above_head', 'notool_left_bend_right_up_from_side']
            },
            
            # ==== SHOULDER FLEXION WITH SIDE BEND (wrist-shoulder-hip for diagonal) ====
            {
                'name': 'R_Wrist_Shoulder_Hip',
                'display': 'Right Diagonal Raise',
                'joints': ('R_wrist', 'R_shoulder', 'R_hip'),
                'instruction': 'Raise your RIGHT hand diagonally up and across',
                'rest_instruction': 'Lower your RIGHT hand down',
                'normal_max': 135,
                'normal_min': 80,
                'used_in': ['notool_raising_hands_diagonally']
            },
            {
                'name': 'L_Wrist_Shoulder_Hip',
                'display': 'Left Diagonal Raise',
                'joints': ('L_wrist', 'L_shoulder', 'L_hip'),
                'instruction': 'Raise your LEFT hand diagonally up and across',
                'rest_instruction': 'Lower your LEFT hand down',
                'normal_max': 135,
                'normal_min': 80,
                'used_in': ['notool_raising_hands_diagonally']
            },
        ]
        
        # Now create Excel file if needed (after calibration_measurements is defined)
        self.ensure_excel_exists()
        
        # Mapping from calibration measurement names to robot demo functions
        self.robot_demo_mapping = {
            'R_Shoulder_Hip_Elbow': ('rom_demo_shoulder_forward_raise', 'right', 'shoulder_forward'),
            'L_Shoulder_Hip_Elbow': ('rom_demo_shoulder_forward_raise', 'left', 'shoulder_forward'),
            'R_Shoulder_Hip_Wrist': ('rom_demo_shoulder_side_raise', 'right', 'shoulder_side'),
            'L_Shoulder_Hip_Wrist': ('rom_demo_shoulder_side_raise', 'left', 'shoulder_side'),
            'R_Elbow': ('rom_demo_elbow_bend', 'right', 'elbow_bend'),
            'L_Elbow': ('rom_demo_elbow_bend', 'left', 'elbow_bend'),
            'R_Wrist_Shoulder_Shoulder': ('rom_demo_arm_across_body', 'right', 'arm_across'),
            'L_Wrist_Shoulder_Shoulder': ('rom_demo_arm_across_body', 'left', 'arm_across'),
            # Add more mappings as needed
        }
    
    def play_calibration_audio(self, audio_key):
        """Play calibration audio with fallback if file doesn't exist"""
        try:
            if audio_key in CALIBRATION_AUDIO:
                say(CALIBRATION_AUDIO[audio_key])
                return get_wav_duration(CALIBRATION_AUDIO[audio_key])
        except:
            # Audio file doesn't exist yet - continue silently
            pass
        return 0
    
    def request_robot_demo(self, measurement_name):
        """Request robot to demonstrate a calibration movement (optional - works without robot too)"""
        # Check if robot is available and running
        robot_available = hasattr(s, 'robot') and s.robot is not None
        
        if not robot_available:
            # No robot - skip demonstration, patient performs based on instructions
            print(f"   üìù {measurement_name} - Follow the text instructions")
            time.sleep(2)
            return None
        
        if measurement_name in self.robot_demo_mapping:
            demo_func, side, audio_key = self.robot_demo_mapping[measurement_name]
            print(f"   ü§ñ Robot demonstrating {measurement_name}...")
            
            # Play audio instruction for this movement type
            audio_duration = self.play_calibration_audio(audio_key)
            
            try:
                # Request robot demonstration using Settings pattern
                s.rom_demo_requested = demo_func
                s.rom_demo_side = side
                s.rom_demo_done = False
                
                # Wait for robot to finish (with timeout)
                timeout = 10  # seconds
                start_time = time.time()
                while not s.rom_demo_done and (time.time() - start_time) < timeout:
                    time.sleep(0.1)
                
                if not s.rom_demo_done:
                    print(f"   ‚ö†Ô∏è Robot demo timeout - continuing without demo")
                
                # Reset flags
                s.rom_demo_requested = None
                s.rom_demo_done = False
                time.sleep(1)  # Brief pause after demo
                return audio_key
            
            except Exception as e:
                print(f"   ‚ö†Ô∏è Robot demo error: {e} - continuing without demo")
                time.sleep(1)
                return None
        else:
            # No specific demo for this measurement, just continue
            print(f"   üìù No robot demo for {measurement_name}, patient will perform independently")
            time.sleep(2)
            return None
    
    def ensure_excel_exists(self):
        """Create Excel file if it doesn't exist"""
        if not os.path.exists(self.EXCEL_FILE):
            print(f"üìÇ Creating ROM calibration database: {self.EXCEL_FILE}")
            self.create_excel_file()
        else:
            print(f"‚úÖ ROM database found: {self.EXCEL_FILE}")
    
    def create_excel_file(self):
        """Create Excel file with proper structure"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Calibration_Data"
        
        # Headers
        headers = ['Patient_ID', 'Calibration_Date', 'Calibration_Time']
        
        # Add column for each measurement (max and min)
        for measurement in self.calibration_measurements:
            headers.append(f"{measurement['name']}_Max")
            headers.append(f"{measurement['name']}_Min")
        
        # Add summary columns
        headers.extend(['Overall_ROM_Score', 'Asymmetry_Score', 'Notes'])
        
        ws.append(headers)
        wb.save(self.EXCEL_FILE)
        print(f"‚úÖ Created {self.EXCEL_FILE} with {len(headers)} columns")
    
    def run_calibration(self):
        """
        Run complete ROM calibration
        Returns True if successful
        """
        print("\n" + "="*70)
        print(f"üéØ COMPREHENSIVE ROM CALIBRATION")
        print(f"   Patient ID: {s.chosen_patient_ID}")
        print(f"   Measuring: {len(self.calibration_measurements)} different movements")
        print("="*70)
        
        # Play welcome audio
        self.play_calibration_audio('start')  # "Let's measure your range of motion"
        time.sleep(2)
        
        # Reset state
        s.patient_calibrated = False
        rom_data = {}
        
        # Run all measurements
        for idx, measurement in enumerate(self.calibration_measurements, 1):
            if s.stop_requested:
                print("\n‚ö†Ô∏è Calibration cancelled by user")
                return False
            
            # Update GUI status
            s.current_calibration_movement = measurement['display']
            s.current_calibration_progress = f"{idx}/{len(self.calibration_measurements)}"
            
            print(f"\n[{idx}/{len(self.calibration_measurements)}] {measurement['display']}")
            
            success, max_val, min_val = self.measure_rom(measurement)
            
            if success:
                rom_data[f"{measurement['name']}_Max"] = max_val
                rom_data[f"{measurement['name']}_Min"] = min_val
                print(f"   ‚úÖ Recorded: {min_val:.1f}¬∞ - {max_val:.1f}¬∞ (Range: {max_val-min_val:.1f}¬∞)")
            else:
                # Use default values
                rom_data[f"{measurement['name']}_Max"] = measurement['normal_max']
                rom_data[f"{measurement['name']}_Min"] = measurement['normal_min']
                print(f"   ‚ö†Ô∏è Using defaults: {measurement['normal_min']}-{measurement['normal_max']}¬∞")
            
            time.sleep(1)  # Pause between measurements
        
        # Calculate scores
        overall_score = self.calculate_rom_score(rom_data)
        asymmetry_score = self.calculate_asymmetry(rom_data)
        
        # Save to Excel
        self.save_to_excel(s.chosen_patient_ID, rom_data, overall_score, asymmetry_score)
        
        # Store in Settings
        s.patient_rom = rom_data
        s.patient_calibrated = True
        
        # Print summary
        print("\n" + "="*70)
        print("‚úÖ CALIBRATION COMPLETE!")
        print("="*70)
        print(f"üìä Overall ROM Score: {overall_score:.1f}/100")
        print(f"üìä Asymmetry Score: {asymmetry_score:.1f}¬∞ (Left vs Right difference)")
        print(f"üíæ Data saved to: {self.EXCEL_FILE}")
        print("="*70)
        
        # Play completion audio
        self.play_calibration_audio('complete')  # "Calibration complete!"
        
        return True
    
    def measure_rom(self, config):
        """
        Measure a single ROM
        Returns (success, max_angle, min_angle)
        """
        joints = config['joints']
        
        # Request robot demonstration (includes audio)
        audio_key = self.request_robot_demo(config['name'])
        
        print(f"   üìù {config['instruction']}")
        print(f"   ü§ñ Watch the robot, then copy the movement...")
        print(f"   ‚è≥ You have 8 seconds to get into position...")
        time.sleep(8)  # Give patient more time to read, understand, and position
        
        # Capture MAXIMUM
        print(f"   üì∏ Capturing maximum angle... Hold the position!")
        self.play_calibration_audio('hold_position')  # "Hold this position"
        max_angle = self.capture_angle(joints, duration=6, target='max')
        
        if max_angle is None:
            return (False, None, None)
        
        print(f"   ‚úÖ Maximum captured: {max_angle:.1f}¬∞")
        time.sleep(2)  # Brief pause
        
        # Capture MINIMUM
        print(f"   üìù {config['rest_instruction']}")
        self.play_calibration_audio('return_rest')  # "Return to rest position"
        print(f"   ‚è≥ You have 8 seconds to get into position...")
        time.sleep(8)  # Give patient time to return to rest position
        
        print(f"   üì∏ Capturing minimum angle... Hold the position!")
        self.play_calibration_audio('hold_position')  # "Hold this position"
        min_angle = self.capture_angle(joints, duration=6, target='min')
        
        if min_angle is None:
            return (False, None, None)
        
        print(f"   ‚úÖ Minimum captured: {min_angle:.1f}¬∞")
        self.play_calibration_audio('good_job')  # "Good job"
        
        # Validate (max should be > min)
        if max_angle < min_angle:
            max_angle, min_angle = min_angle, max_angle
        
        print(f"   ‚è∏Ô∏è  Relax for 3 seconds before next measurement...")
        time.sleep(3)  # Give patient time to relax between measurements
        
        return (True, max_angle, min_angle)
    
    def capture_angle(self, joints, duration=4, target='max'):
        """Capture angles from camera"""
        samples = []
        start_time = time.time()
        
        while time.time() - start_time < duration:
            if s.stop_requested:
                return None
            
            skeleton = s.camera.get_skeleton_data()
            
            if skeleton is not None:
                try:
                    j1 = skeleton[joints[0]]
                    j2 = skeleton[joints[1]]
                    j3 = skeleton[joints[2]]
                    
                    angle = s.camera.calc_angle_3d(j1, j2, j3, "calibration")
                    
                    if angle is not None and angle > 0:
                        samples.append(angle)
                
                except (KeyError, AttributeError):
                    pass
            
            time.sleep(0.1)
        
        if samples:
            result = np.max(samples) if target == 'max' else np.min(samples)
            return result
        
        return None
    
    def save_to_excel(self, patient_id, rom_data, overall_score, asymmetry_score):
        """Save calibration data to Excel"""
        try:
            wb = load_workbook(self.EXCEL_FILE)
            ws = wb["Calibration_Data"]
            
            # Check if patient already has calibration
            df = pd.read_excel(self.EXCEL_FILE)
            existing = df[df['Patient_ID'].astype(str) == str(patient_id)]
            
            if not existing.empty:
                # Update existing row
                row_idx = existing.index[0] + 2  # +2 for header and 0-indexing
                print(f"   üìù Updating existing calibration for patient {patient_id}")
            else:
                # Add new row
                row_idx = ws.max_row + 1
                print(f"   üìù Adding new calibration for patient {patient_id}")
            
            # Prepare data
            now = datetime.now()
            ws.cell(row_idx, 1, str(patient_id))
            ws.cell(row_idx, 2, now.strftime("%Y-%m-%d"))
            ws.cell(row_idx, 3, now.strftime("%H:%M:%S"))
            
            # Add ROM data
            col = 4
            for measurement in self.calibration_measurements:
                ws.cell(row_idx, col, rom_data.get(f"{measurement['name']}_Max", measurement['normal_max']))
                ws.cell(row_idx, col+1, rom_data.get(f"{measurement['name']}_Min", measurement['normal_min']))
                col += 2
            
            # Add scores
            ws.cell(row_idx, col, overall_score)
            ws.cell(row_idx, col+1, asymmetry_score)
            ws.cell(row_idx, col+2, "Initial calibration")
            
            wb.save(self.EXCEL_FILE)
            print(f"   ‚úÖ Saved to Excel successfully!")
            
        except Exception as e:
            print(f"   ‚ùå Error saving to Excel: {e}")
    
    def load_from_excel(self, patient_id):
        """
        Load patient's ROM from Excel
        Returns dictionary with ROM data or None
        """
        try:
            df = pd.read_excel(self.EXCEL_FILE)
            patient_data = df[df['Patient_ID'].astype(str) == str(patient_id)]
            
            if patient_data.empty:
                print(f"‚ö†Ô∏è No calibration found for patient {patient_id}")
                return None
            
            # Convert to dictionary
            rom_dict = patient_data.iloc[0].to_dict()
            
            print(f"‚úÖ Loaded calibration for patient {patient_id}")
            return rom_dict
            
        except Exception as e:
            print(f"‚ùå Error loading from Excel: {e}")
            return None
    
    def calculate_rom_score(self, rom_data):
        """Calculate overall ROM score (0-100)"""
        scores = []
        
        for measurement in self.calibration_measurements:
            max_key = f"{measurement['name']}_Max"
            if max_key in rom_data:
                patient_max = rom_data[max_key]
                normal_max = measurement['normal_max']
                score = min(100, (patient_max / normal_max) * 100)
                scores.append(score)
        
        return np.mean(scores) if scores else 0
    
    def calculate_asymmetry(self, rom_data):
        """Calculate left-right asymmetry"""
        asymmetries = []
        
        # Compare R vs L for each measurement type
        pairs = [
            ('R_Shoulder_Hip_Elbow', 'L_Shoulder_Hip_Elbow'),
            ('R_Elbow', 'L_Elbow'),
            ('R_Wrist_Shoulder_Hip', 'L_Wrist_Shoulder_Hip'),
        ]
        
        for r_name, l_name in pairs:
            r_max = rom_data.get(f"{r_name}_Max", 0)
            l_max = rom_data.get(f"{l_name}_Max", 0)
            if r_max > 0 and l_max > 0:
                asymmetries.append(abs(r_max - l_max))
        
        return np.mean(asymmetries) if asymmetries else 0


# ==================== HELPER FUNCTIONS ====================

def load_patient_rom_on_start(patient_id):
    """
    Load patient's ROM when training starts
    Call this at the beginning of training
    """
    calibration = Patient_Calibration()
    rom_data = calibration.load_from_excel(patient_id)
    
    if rom_data:
        s.patient_rom = rom_data
        s.patient_calibrated = True
        return True
    else:
        s.patient_calibrated = False
        s.patient_rom = {}
        return False


def get_adaptive_range_for_joint(joint_combo, default_min, default_max):
    """
    Get adaptive range based on patient's ROM
    
    Args:
        joint_combo: e.g., 'R_Elbow', 'L_Shoulder_Hip_Elbow', etc.
        default_min, default_max: Fallback values
    
    Returns:
        (adjusted_min, adjusted_max)
    """
    if not hasattr(s, 'patient_calibrated') or not s.patient_calibrated:
        return (default_min, default_max)
    
    if not hasattr(s, 'patient_rom') or not s.patient_rom:
        return (default_min, default_max)
    
    # Get patient's max and min for this joint combo
    max_key = f"{joint_combo}_Max"
    min_key = f"{joint_combo}_Min"
    
    if max_key not in s.patient_rom or min_key not in s.patient_rom:
        return (default_min, default_max)
    
    patient_max = s.patient_rom[max_key]
    patient_min = s.patient_rom[min_key]
    
    # Safety factor
    safety = 0.90
    range_span = patient_max - patient_min
    safe_max = patient_min + (range_span * safety)
    
    # If patient is limited, use their range
    if patient_max < default_max:
        return (patient_min, safe_max)
    
    return (default_min, default_max)


# ==================== TESTING ====================

if __name__ == "__main__":
    print("="*70)
    print("üè• Patient Calibration - Test Mode")
    print("="*70)
    print(f"\nüìä This system measures {len(Patient_Calibration().calibration_measurements)} movements")
    print("\n‚úÖ Module loaded successfully!")

